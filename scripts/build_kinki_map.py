from __future__ import annotations

import html
import json
import math
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = Path(r"C:\Users\AMALINK-PC\Downloads\大阪市役所起点_出張距離一覧 (1).xlsx")
GEOJSON = ROOT / "prefectures.geojson"
OUTPUT_DIR = ROOT / "outputs" / "kinki_map"
PNG_OUT = OUTPUT_DIR / "kinki_business_trip_pin_map_labeled.png"
HTML_OUT = OUTPUT_DIR / "kinki_business_trip_pin_map_labeled.html"
SVG_OUT = OUTPUT_DIR / "kinki_business_trip_pin_map_labeled.svg"

WIDTH = 2600
HEIGHT = 2000

KINKI_PREFS = {"大阪府", "京都府", "兵庫県", "奈良県", "滋賀県", "和歌山県"}

STYLE = {
    "hotel_only": {
        "label": "ホテルのみ",
        "match": {"ホテル代支給"},
        "color": "#2563eb",
        "outline": "#0f2f73",
    },
    "hotel_incentive": {
        "label": "ホテル+出張インセンティブ",
        "match": {"ホテル代＋出張インセンティブ", "ホテル代+出張インセンティブ"},
        "color": "#dc2626",
        "outline": "#7f1d1d",
    },
    "none": {
        "label": "該当なし",
        "match": {"該当なし", "出発点"},
        "color": "#7a7f87",
        "outline": "#3f444c",
    },
}

PREF_LABELS = {
    "兵庫県": (134.90, 35.05),
    "京都府": (135.52, 35.25),
    "滋賀県": (136.13, 35.22),
    "大阪府": (135.48, 34.60),
    "奈良県": (135.85, 34.38),
    "和歌山県": (135.36, 33.96),
}


def category(value: str) -> str:
    text = str(value or "").strip()
    for key, meta in STYLE.items():
        if text in meta["match"]:
            return key
    if "インセンティブ" in text:
        return "hotel_incentive"
    if "ホテル" in text:
        return "hotel_only"
    return "none"


def load_rows() -> list[dict]:
    wb = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    ws = wb["全役場一覧"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("都道府県") not in KINKI_PREFS:
            continue
        if rec.get("緯度") is None or rec.get("経度") is None:
            continue
        rec["category"] = category(str(rec.get("出張区分", "")))
        rows.append(rec)
    return rows


def lonlat_iter(coords):
    if isinstance(coords[0], (float, int)):
        yield float(coords[0]), float(coords[1])
    else:
        for child in coords:
            yield from lonlat_iter(child)


def load_prefectures() -> list[dict]:
    data = json.loads(GEOJSON.read_text(encoding="utf-8"))
    return [f for f in data["features"] if f["properties"]["P"] in KINKI_PREFS]


def mercator(lon: float, lat: float) -> tuple[float, float]:
    lat = max(min(lat, 85.0), -85.0)
    x = math.radians(lon)
    y = math.log(math.tan(math.pi / 4 + math.radians(lat) / 2))
    return x, y


def build_projection(features: list[dict], rows: list[dict], width: int, height: int):
    xs: list[float] = []
    ys: list[float] = []
    for feature in features:
        for lon, lat in lonlat_iter(feature["geometry"]["coordinates"]):
            x, y = mercator(lon, lat)
            xs.append(x)
            ys.append(y)
    for row in rows:
        x, y = mercator(float(row["経度"]), float(row["緯度"]))
        xs.append(x)
        ys.append(y)

    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    dx = max_x - min_x
    dy = max_y - min_y
    min_x -= dx * 0.05
    max_x += dx * 0.05
    min_y -= dy * 0.05
    max_y += dy * 0.05

    left, right, top, bottom = 62, 62, 86, 78
    map_w = width - left - right
    map_h = height - top - bottom
    scale = min(map_w / (max_x - min_x), map_h / (max_y - min_y))
    used_w = (max_x - min_x) * scale
    used_h = (max_y - min_y) * scale
    ox = left + (map_w - used_w) / 2
    oy = top + (map_h - used_h) / 2

    def project(lon: float, lat: float) -> tuple[float, float]:
        x, y = mercator(lon, lat)
        px = ox + (x - min_x) * scale
        py = oy + (max_y - y) * scale
        return px, py

    return project


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        r"C:\Windows\Fonts\meiryob.ttc" if bold else r"C:\Windows\Fonts\meiryo.ttc",
        r"C:\Windows\Fonts\YuGothB.ttc" if bold else r"C:\Windows\Fonts\YuGothR.ttc",
        r"C:\Windows\Fonts\arial.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def polygon_rings(geometry: dict):
    coords = geometry["coordinates"]
    if geometry["type"] == "Polygon":
        for polygon in [coords]:
            yield polygon
    elif geometry["type"] == "MultiPolygon":
        for polygon in coords:
            yield polygon


def draw_pin(draw: ImageDraw.ImageDraw, x: float, y: float, color: str, outline: str, r: int = 4):
    draw.ellipse((x - r, y - r, x + r, y + r), fill=color, outline=outline, width=2)
    draw.line((x, y + r - 1, x, y + r + 5), fill=outline, width=2)


def text_size(draw: ImageDraw.ImageDraw, text: str, font_obj):
    box = draw.textbbox((0, 0), text, font=font_obj)
    return box[2] - box[0], box[3] - box[1]


def place_name(name: str) -> str:
    text = str(name or "")
    for suffix in ("市役所", "町役場", "村役場", "区役所", "役所", "役場"):
        if text.endswith(suffix):
            return text[: -len(suffix)] + suffix[0]
    return text


def intersects(a, b) -> bool:
    return not (a[2] < b[0] or b[2] < a[0] or a[3] < b[1] or b[3] < a[1])


def draw_text_halo(draw: ImageDraw.ImageDraw, xy, text: str, font_obj, fill="#111827"):
    x, y = xy
    for dx, dy in [(-2, 0), (2, 0), (0, -2), (0, 2), (-1, -1), (1, -1), (-1, 1), (1, 1)]:
        draw.text((x + dx, y + dy), text, fill="#ffffff", font=font_obj)
    draw.text((x, y), text, fill=fill, font=font_obj)


def render_png(features: list[dict], rows: list[dict], project):
    width, height = WIDTH, HEIGHT
    image = Image.new("RGB", (width, height), "#eef6fb")
    draw = ImageDraw.Draw(image)

    title_font = font(48, True)
    regular_font = font(31)
    small_font = font(25)
    pin_label_font = font(13, True)
    label_font = font(36, True)

    draw.rectangle((0, 0, width, 112), fill="#ffffff")
    draw.text((56, 28), "近畿二府四県 出張区分ピンマップ", fill="#111827", font=title_font)
    draw.text((56, height - 64), "座標: スプレッドシートの緯度・経度列 / 境界: 都道府県GeoJSON", fill="#374151", font=small_font)

    for feature in features:
        for polygon in polygon_rings(feature["geometry"]):
            for ring_index, ring in enumerate(polygon):
                pts = [project(lon, lat) for lon, lat in ring]
                if len(pts) >= 3:
                    fill = "#f8fbf8" if ring_index == 0 else "#eef6fb"
                    draw.polygon(pts, fill=fill)
                    draw.line(pts + [pts[0]], fill="#aab6c2", width=2)

    for pref, (lon, lat) in PREF_LABELS.items():
        x, y = project(lon, lat)
        w, h = text_size(draw, pref, label_font)
        draw.text((x - w / 2, y - h / 2), pref, fill="#64748b", font=label_font)

    rows_sorted = sorted(rows, key=lambda r: 0 if r["category"] == "none" else 1)
    for row in rows_sorted:
        x, y = project(float(row["経度"]), float(row["緯度"]))
        meta = STYLE[row["category"]]
        draw_pin(draw, x, y, meta["color"], meta["outline"], 4)

    label_boxes = []
    for row in sorted(rows, key=lambda r: (r["category"] == "none", -float(r["距離 (km)"] or 0))):
        x, y = project(float(row["経度"]), float(row["緯度"]))
        label = place_name(row["役場名"])
        tw, th = text_size(draw, label, pin_label_font)
        candidates = [
            (x + 7, y - th - 2),
            (x + 7, y + 3),
            (x - tw - 7, y - th - 2),
            (x - tw - 7, y + 3),
            (x - tw / 2, y - th - 11),
            (x - tw / 2, y + 10),
        ]
        chosen = None
        for cx, cy in candidates:
            box = (cx - 2, cy - 2, cx + tw + 2, cy + th + 2)
            if box[0] < 8 or box[1] < 116 or box[2] > width - 8 or box[3] > height - 70:
                continue
            if not any(intersects(box, old) for old in label_boxes):
                chosen = (cx, cy, box)
                break
        if chosen is None:
            cx, cy = candidates[0]
            box = (cx - 2, cy - 2, cx + tw + 2, cy + th + 2)
            chosen = (cx, cy, box)
        label_boxes.append(chosen[2])
        draw_text_halo(draw, (chosen[0], chosen[1]), label, pin_label_font, "#111827")

    # Osaka City Hall origin gets a subtle ring while retaining the gray category color.
    origin = next((r for r in rows if r.get("出張区分") == "出発点"), None)
    if origin:
        x, y = project(float(origin["経度"]), float(origin["緯度"]))
        draw.ellipse((x - 11, y - 11, x + 11, y + 11), outline="#111827", width=3)

    legend_x, legend_y = width - 650, 136
    draw.rounded_rectangle((legend_x, legend_y, width - 56, legend_y + 240), radius=8, fill="#ffffff", outline="#d6dee8", width=2)
    draw.text((legend_x + 32, legend_y + 26), "出張区分", fill="#111827", font=regular_font)
    counts = Counter(r["category"] for r in rows)
    y = legend_y + 82
    for key in ["hotel_incentive", "hotel_only", "none"]:
        meta = STYLE[key]
        draw_pin(draw, legend_x + 44, y + 13, meta["color"], meta["outline"], 6)
        draw.text((legend_x + 78, y), f"{meta['label']}  {counts[key]}件", fill="#111827", font=small_font)
        y += 50

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    image.save(PNG_OUT)


def svg_path_for_ring(ring, project) -> str:
    points = [project(lon, lat) for lon, lat in ring]
    if not points:
        return ""
    start = f"M {points[0][0]:.2f} {points[0][1]:.2f}"
    rest = " ".join(f"L {x:.2f} {y:.2f}" for x, y in points[1:])
    return f"{start} {rest} Z"


def render_svg_html(features: list[dict], rows: list[dict], project):
    width, height = WIDTH, HEIGHT
    counts = Counter(r["category"] for r in rows)
    paths = []
    for feature in features:
        pref = feature["properties"]["P"]
        for polygon in polygon_rings(feature["geometry"]):
            for ring in polygon:
                d = svg_path_for_ring(ring, project)
                if d:
                    paths.append(f'<path d="{d}" class="pref"><title>{html.escape(pref)}</title></path>')

    labels = []
    for pref, (lon, lat) in PREF_LABELS.items():
        x, y = project(lon, lat)
        labels.append(f'<text x="{x:.1f}" y="{y:.1f}" class="pref-label">{html.escape(pref)}</text>')

    pins = []
    for row in rows:
        x, y = project(float(row["経度"]), float(row["緯度"]))
        meta = STYLE[row["category"]]
        tip = f'{row["役場名"]} / {row["都道府県"]} / {row["出張区分"]} / {row["距離 (km)"]}km'
        label = place_name(row["役場名"])
        pins.append(
            f'<g class="pin {row["category"]}" transform="translate({x:.2f},{y:.2f})">'
            f'<line y1="5" y2="14" stroke="{meta["outline"]}" stroke-width="2"/>'
            f'<circle r="4" fill="{meta["color"]}" stroke="{meta["outline"]}" stroke-width="2"/>'
            f'<text x="8" y="-6" class="pin-label">{html.escape(label)}</text>'
            f'<title>{html.escape(tip)}</title></g>'
        )

    legend_items = []
    y = 152
    for key in ["hotel_incentive", "hotel_only", "none"]:
        meta = STYLE[key]
        legend_items.append(
            f'<g><circle cx="2024" cy="{y}" r="9" fill="{meta["color"]}" stroke="{meta["outline"]}" stroke-width="2"/>'
            f'<text x="2054" y="{y + 8}" class="legend-text">{html.escape(meta["label"])} {counts[key]}件</text></g>'
        )
        y += 52

    svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}" role="img" aria-label="近畿二府四県 出張区分ピンマップ">
  <style>
    .title {{ font: 700 48px "Meiryo", "Yu Gothic", sans-serif; fill: #111827; }}
    .note {{ font: 25px "Meiryo", "Yu Gothic", sans-serif; fill: #374151; }}
    .pref {{ fill: #f8fbf8; stroke: #aab6c2; stroke-width: 2; vector-effect: non-scaling-stroke; }}
    .pref-label {{ font: 700 36px "Meiryo", "Yu Gothic", sans-serif; fill: #64748b; text-anchor: middle; dominant-baseline: central; pointer-events: none; }}
    .pin-label {{ font: 700 13px "Meiryo", "Yu Gothic", sans-serif; fill: #111827; stroke: #fff; stroke-width: 4; paint-order: stroke; }}
    .legend-title {{ font: 31px "Meiryo", "Yu Gothic", sans-serif; fill: #111827; }}
    .legend-text {{ font: 25px "Meiryo", "Yu Gothic", sans-serif; fill: #111827; }}
    .pin {{ cursor: help; opacity: .93; }}
    .pin:hover circle {{ r: 10; stroke-width: 3; }}
  </style>
  <rect width="100%" height="100%" fill="#eef6fb"/>
  <rect x="0" y="0" width="{width}" height="112" fill="#fff"/>
  <text x="56" y="76" class="title">近畿二府四県 出張区分ピンマップ</text>
  <g>{''.join(paths)}</g>
  <g>{''.join(labels)}</g>
  <g>{''.join(pins)}</g>
  <rect x="1960" y="136" width="584" height="240" rx="8" fill="#fff" stroke="#d6dee8" stroke-width="2"/>
  <text x="1992" y="190" class="legend-title">出張区分</text>
  <g>{''.join(legend_items)}</g>
  <text x="56" y="1936" class="note">座標: スプレッドシートの緯度・経度列 / 境界: 都道府県GeoJSON</text>
</svg>'''

    html_doc = f'''<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>近畿二府四県 出張区分ピンマップ</title>
  <style>
    body {{ margin: 0; background: #eef6fb; font-family: Meiryo, "Yu Gothic", sans-serif; }}
    main {{ max-width: 1500px; margin: 0 auto; }}
    svg {{ display: block; width: 100%; height: auto; }}
  </style>
</head>
<body><main>{svg}</main></body>
</html>'''

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SVG_OUT.write_text(svg, encoding="utf-8")
    HTML_OUT.write_text(html_doc, encoding="utf-8")


def main():
    rows = load_rows()
    features = load_prefectures()
    project = build_projection(features, rows, WIDTH, HEIGHT)
    render_png(features, rows, project)
    render_svg_html(features, rows, project)
    counts = Counter(r["category"] for r in rows)
    print(f"rows={len(rows)}")
    print(dict(counts))
    print(PNG_OUT)
    print(HTML_OUT)
    print(SVG_OUT)


if __name__ == "__main__":
    main()
