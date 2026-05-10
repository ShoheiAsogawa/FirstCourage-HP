from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
GEOJSON = ROOT / "prefectures.geojson"
OUT = ROOT / "trip-map-system" / "map-data.js"
VIEW_W = 1600
VIEW_H = 1200

KINKI_PREFS = {
    "\u5927\u962a\u5e9c",
    "\u4eac\u90fd\u5e9c",
    "\u5175\u5eab\u770c",
    "\u5948\u826f\u770c",
    "\u6ecb\u8cc0\u770c",
    "\u548c\u6b4c\u5c71\u770c",
    "\u4e09\u91cd\u770c",
}

PREF_LABELS = {
    "\u5175\u5eab\u770c": (134.90, 35.05),
    "\u4eac\u90fd\u5e9c": (135.52, 35.25),
    "\u6ecb\u8cc0\u770c": (136.13, 35.22),
    "\u5927\u962a\u5e9c": (135.48, 34.60),
    "\u5948\u826f\u770c": (135.85, 34.38),
    "\u548c\u6b4c\u5c71\u770c": (135.36, 33.96),
    "\u4e09\u91cd\u770c": (136.45, 34.55),
}


def latest_input() -> Path:
    downloads = Path.home() / "Downloads"
    candidates = sorted(downloads.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        if "(2)" in path.name:
            return path
    for path in candidates:
        if "(1)" in path.name:
            return path
    if candidates:
        return candidates[0]
    raise FileNotFoundError("No xlsx files found in Downloads")


def category_key(text: str) -> str:
    if "\u30a4\u30f3\u30bb\u30f3\u30c6\u30a3\u30d6" in text:
        return "hotel_incentive"
    if "\u30db\u30c6\u30eb" in text:
        return "hotel_only"
    return "none"


def municipality_name(office_name: str) -> str:
    text = str(office_name)
    for suffix in ("\u5e02\u5f79\u6240", "\u753a\u5f79\u5834", "\u6751\u5f79\u5834", "\u533a\u5f79\u6240", "\u5f79\u6240", "\u5f79\u5834"):
        if text.endswith(suffix):
            return text[: -len(suffix)] + suffix[0]
    return text


def load_rows() -> list[dict]:
    wb = load_workbook(latest_input(), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        pref, office, lat, lon, distance, travel_time, trip_class = values[:7]
        if pref not in KINKI_PREFS or lat is None or lon is None:
            continue
        rows.append(
            {
                "id": f"p{idx}",
                "prefecture": str(pref),
                "officeName": str(office),
                "municipality": municipality_name(str(office)),
                "lat": round(float(lat), 6),
                "lon": round(float(lon), 6),
                "distanceKm": round(float(distance or 0), 1),
                "travelMinutes": int(travel_time or 0),
                "tripClass": str(trip_class),
                "category": category_key(str(trip_class)),
            }
        )
    return rows


def lonlat_iter(coords):
    if isinstance(coords[0], (float, int)):
        yield float(coords[0]), float(coords[1])
    else:
        for child in coords:
            yield from lonlat_iter(child)


def load_features() -> list[dict]:
    data = json.loads(GEOJSON.read_text(encoding="utf-8"))
    return [feature for feature in data["features"] if feature["properties"]["P"] in KINKI_PREFS]


def mercator(lon: float, lat: float) -> tuple[float, float]:
    lat = max(min(lat, 85), -85)
    return math.radians(lon), math.log(math.tan(math.pi / 4 + math.radians(lat) / 2))


def build_project(features: list[dict], rows: list[dict]):
    xs, ys = [], []
    for feature in features:
        for lon, lat in lonlat_iter(feature["geometry"]["coordinates"]):
            x, y = mercator(lon, lat)
            xs.append(x)
            ys.append(y)
    for row in rows:
        x, y = mercator(row["lon"], row["lat"])
        xs.append(x)
        ys.append(y)

    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    dx, dy = max_x - min_x, max_y - min_y
    min_x -= dx * 0.045
    max_x += dx * 0.045
    min_y -= dy * 0.045
    max_y += dy * 0.045

    left, right, top, bottom = 34, 34, 48, 44
    scale = min((VIEW_W - left - right) / (max_x - min_x), (VIEW_H - top - bottom) / (max_y - min_y))
    used_w = (max_x - min_x) * scale
    used_h = (max_y - min_y) * scale
    ox = left + ((VIEW_W - left - right) - used_w) / 2
    oy = top + ((VIEW_H - top - bottom) - used_h) / 2

    def project(lon: float, lat: float) -> tuple[float, float]:
        x, y = mercator(lon, lat)
        return ox + (x - min_x) * scale, oy + (max_y - y) * scale

    return project


def perpendicular_distance(point, start, end) -> float:
    px, py = point
    sx, sy = start
    ex, ey = end
    dx, dy = ex - sx, ey - sy
    if dx == 0 and dy == 0:
        return math.hypot(px - sx, py - sy)
    return abs(dy * px - dx * py + ex * sy - ey * sx) / math.hypot(dx, dy)


def simplify(points: list[tuple[float, float]], tolerance: float = 0.75) -> list[tuple[float, float]]:
    if len(points) <= 3:
        return points
    start, end = points[0], points[-1]
    max_dist, index = 0.0, 0
    for i in range(1, len(points) - 1):
        dist = perpendicular_distance(points[i], start, end)
        if dist > max_dist:
            max_dist, index = dist, i
    if max_dist > tolerance:
        left = simplify(points[: index + 1], tolerance)
        right = simplify(points[index:], tolerance)
        return left[:-1] + right
    return [start, end]


def polygon_rings(geometry: dict):
    if geometry["type"] == "Polygon":
        yield from geometry["coordinates"]
    elif geometry["type"] == "MultiPolygon":
        for polygon in geometry["coordinates"]:
            yield from polygon


def path_d(points: list[tuple[float, float]]) -> str:
    if not points:
        return ""
    head = f"M {points[0][0]:.1f} {points[0][1]:.1f}"
    tail = " ".join(f"L {x:.1f} {y:.1f}" for x, y in points[1:])
    return f"{head} {tail} Z"


def main() -> None:
    rows = load_rows()
    features = load_features()
    project = build_project(features, rows)

    paths = []
    for feature in features:
        pref = feature["properties"]["P"]
        for ring in polygon_rings(feature["geometry"]):
            d = path_d(simplify([project(lon, lat) for lon, lat in ring], 0.75))
            if d:
                paths.append({"prefecture": pref, "d": d})

    labels = []
    for pref, (lon, lat) in PREF_LABELS.items():
        x, y = project(lon, lat)
        labels.append({"prefecture": pref, "x": round(x, 1), "y": round(y, 1)})

    pins = []
    for row in rows:
        x, y = project(row["lon"], row["lat"])
        pins.append({**row, "x": round(x, 1), "y": round(y, 1)})

    payload = {"viewBox": [0, 0, VIEW_W, VIEW_H], "paths": paths, "prefectureLabels": labels, "pins": pins}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "window.TRIP_MAP_DATA="
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(f"wrote {OUT}")
    print(f"pins={len(pins)} paths={len(paths)}")


if __name__ == "__main__":
    main()
